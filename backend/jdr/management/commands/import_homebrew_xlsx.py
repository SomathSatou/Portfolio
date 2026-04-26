"""
Import Homebrew XLSX catalogue files (inventory + spellbook) into the JDR database.

Usage:
    python manage.py import_homebrew_xlsx \
        --inventory "path/to/Homebrew_inventory.xlsx" \
        --spellbook "path/to/Homebrew_spellbook.xlsx" \
        --campaign 1
"""
import re
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from jdr.models import Campaign, Item, Spell


# ─── Constants ────────────────────────────────────────────────────────────────

RARITY_COLOR_MAP = {
    '#626262': 'commun',
    '#27ae60': 'peu_commun',
    '#2980b9': 'rare',
    '#8e44ad': 'très_rare',
    '#e67e22': 'légendaire',
    '#b60e16': 'artéfact',
}

SHEET_ITEM_TYPE = {
    'weapon': 'Arme',
    'armor': 'Armure',
    'magic': 'Objet magique',
    'other': 'Divers',
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _patch_openpyxl():
    """Monkey-patch openpyxl DataValidation to accept any errorStyle value."""
    from openpyxl.worksheet.datavalidation import DataValidation
    _orig_init = DataValidation.__init__

    def _patched_init(self, *args, **kwargs):
        if 'errorStyle' in kwargs and kwargs['errorStyle'] not in (
            'warning', 'stop', 'information', None,
        ):
            kwargs['errorStyle'] = None
        _orig_init(self, *args, **kwargs)

    DataValidation.__init__ = _patched_init


def strip_html(html) -> str:
    """Remove HTML tags, keeping text content."""
    if not html:
        return ''
    text = str(html)
    text = re.sub(r'<br\s*/?>', '\n', text)
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'\r\n', '\n', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def safe_str(val) -> str:
    """Convert cell value to string, empty if None."""
    if val is None:
        return ''
    return str(val).strip()


def safe_decimal(val) -> Decimal:
    """Convert cell value to Decimal."""
    if val is None:
        return Decimal('0')
    try:
        return Decimal(str(val))
    except Exception:
        return Decimal('0')


class Command(BaseCommand):
    help = 'Import Homebrew XLSX catalogue (inventory + spellbook) into the JDR database.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--inventory', type=str, default='',
            help='Path to Homebrew_inventory.xlsx',
        )
        parser.add_argument(
            '--spellbook', type=str, default='',
            help='Path to Homebrew_spellbook.xlsx',
        )
        parser.add_argument(
            '--campaign', type=int, required=True,
            help='Campaign ID to assign items/spells to',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Preview what would be imported without writing to DB.',
        )

    def handle(self, *args, **options):
        dry_run = options['dry_run']

        try:
            campaign = Campaign.objects.get(pk=options['campaign'])
        except Campaign.DoesNotExist:
            raise CommandError(f"Campaign with id={options['campaign']} not found.")

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no changes will be written.'))

        _patch_openpyxl()
        import openpyxl  # noqa: E402 — import after patch

        inv_path = options['inventory']
        sb_path = options['spellbook']

        if inv_path:
            p = Path(inv_path)
            if not p.exists():
                raise CommandError(f'Inventory file not found: {p}')
            self._import_inventory(openpyxl, p, campaign, dry_run)

        if sb_path:
            p = Path(sb_path)
            if not p.exists():
                raise CommandError(f'Spellbook file not found: {p}')
            self._import_spellbook(openpyxl, p, campaign, dry_run)

        self.stdout.write(self.style.SUCCESS('\nImport XLSX terminé !'))

    # ─── Inventory ────────────────────────────────────────────────────────

    def _import_inventory(self, openpyxl, path: Path, campaign, dry_run: bool):
        wb = openpyxl.load_workbook(str(path))
        total = 0

        for sheet_name in ('weapon', 'armor', 'magic', 'other'):
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet "{sheet_name}" not found, skipping.'))
                continue

            ws = wb[sheet_name]
            item_type = SHEET_ITEM_TYPE[sheet_name]
            created_count = 0
            skipped_count = 0

            self.stdout.write(f'\n── Sheet: {sheet_name} ({item_type}) ──')

            for row in ws.iter_rows(min_row=2, values_only=True):
                name = safe_str(row[1]) if len(row) > 1 else ''
                if not name:
                    continue

                rarity_color = safe_str(row[2]) if len(row) > 2 else '#626262'
                rarity = RARITY_COLOR_MAP.get(rarity_color, 'commun')

                # Columns differ by sheet type
                properties = {}
                if sheet_name == 'weapon':
                    # icon, name, rarity, damage, range, description, price
                    properties['damage'] = safe_str(row[3]) if len(row) > 3 else ''
                    properties['range'] = safe_str(row[4]) if len(row) > 4 else ''
                    description = strip_html(row[5]) if len(row) > 5 else ''
                    price = safe_decimal(row[6]) if len(row) > 6 else Decimal('0')
                elif sheet_name == 'armor':
                    # icon, name, rarity, armorpoint, description, price
                    properties['armor_point'] = safe_str(row[3]) if len(row) > 3 else ''
                    description = strip_html(row[4]) if len(row) > 4 else ''
                    price = safe_decimal(row[5]) if len(row) > 5 else Decimal('0')
                elif sheet_name == 'magic':
                    # icon, name, rarity, charges, description, price
                    properties['charges'] = safe_str(row[3]) if len(row) > 3 else ''
                    description = strip_html(row[4]) if len(row) > 4 else ''
                    price = safe_decimal(row[5]) if len(row) > 5 else Decimal('0')
                else:
                    # other: icon, name, rarity, description, price
                    description = strip_html(row[3]) if len(row) > 3 else ''
                    price = safe_decimal(row[4]) if len(row) > 4 else Decimal('0')

                icon = safe_str(row[0]) if len(row) > 0 else ''
                if icon:
                    properties['icon'] = icon

                is_magical = sheet_name == 'magic'

                if dry_run:
                    self.stdout.write(f'  [DRY] {name} ({rarity}, {price} PO)')
                    created_count += 1
                    continue

                _, created = Item.objects.update_or_create(
                    campaign=campaign, name=name,
                    defaults={
                        'description': description,
                        'rarity': rarity,
                        'item_type': item_type,
                        'value': price,
                        'properties': properties,
                        'is_magical': is_magical,
                    },
                )
                if created:
                    created_count += 1
                else:
                    skipped_count += 1

            total += created_count
            self.stdout.write(
                f'  → {created_count} created, {skipped_count} updated'
            )

        self.stdout.write(f'\nTotal inventory items: {total}')

    # ─── Spellbook ────────────────────────────────────────────────────────

    def _import_spellbook(self, openpyxl, path: Path, campaign, dry_run: bool):
        wb = openpyxl.load_workbook(str(path))

        if 'spell' not in wb.sheetnames:
            raise CommandError('Sheet "spell" not found in spellbook XLSX.')

        ws = wb['spell']
        created_count = 0
        skipped_count = 0

        self.stdout.write(f'\n── Spellbook ──')

        # Headers: name, tag, effect, castingTime, unitCT, range, components,
        #          description, energyCost (mana)
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = safe_str(row[0]) if len(row) > 0 else ''
            if not name:
                continue

            school = safe_str(row[1]) if len(row) > 1 else ''
            damage = safe_str(row[2]) if len(row) > 2 else ''
            casting_time = safe_str(row[3]) if len(row) > 3 else ''
            unit_ct = safe_str(row[4]) if len(row) > 4 else ''
            range_distance = safe_str(row[5]) if len(row) > 5 else ''
            components = safe_str(row[6]) if len(row) > 6 else ''
            description = strip_html(row[7]) if len(row) > 7 else ''
            mana_raw = safe_str(row[8]) if len(row) > 8 else '0'

            try:
                mana_cost = int(float(mana_raw)) if mana_raw else 0
            except (ValueError, TypeError):
                mana_cost = 0

            extra = {}
            if unit_ct:
                extra['unitCT'] = unit_ct
            if components:
                extra['components'] = components

            if dry_run:
                self.stdout.write(f'  [DRY] {name} (mana={mana_cost}, dmg={damage})')
                created_count += 1
                continue

            _, created = Spell.objects.update_or_create(
                campaign=campaign, name=name,
                defaults={
                    'description': description,
                    'mana_cost': mana_cost,
                    'damage': damage,
                    'range_distance': range_distance,
                    'casting_time': casting_time,
                    'school': school,
                    'extra': extra,
                },
            )
            if created:
                created_count += 1
            else:
                skipped_count += 1

        self.stdout.write(f'  → {created_count} created, {skipped_count} updated')
